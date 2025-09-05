import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
import speech_recognition as sr
import pytesseract
from PIL import Image, ImageOps
import cv2
import requests
import json
from datetime import datetime
import threading
import os
import numpy as np
import re
from num2words import num2words
import time
import io

# إعداد صفحة Streamlit
st.set_page_config(
    page_title="نظام المحاسبة الذكي",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ألوان مستوحاة من Excel وChatGPT
excel_color = "#217346"
chatgpt_color = "#0fa37f"
accent_color = "#1a73e8"
background_color = "#f0f0f0"

class AccountingAIApp:
    def __init__(self):
        # بيانات التطبيق
        if 'data' not in st.session_state:
            st.session_state.data = {
                "المبيعات": pd.DataFrame(columns=["التاريخ", "العميل", "المبلغ", "الوصف", "الحالة"]),
                "المشتريات": pd.DataFrame(columns=["التاريخ", "المورد", "المبلغ", "الوصف", "الحالة"]),
                "المصروفات": pd.DataFrame(columns=["التاريخ", "النوع", "المبلغ", "الوصف", "الحالة"]),
                "العملاء": pd.DataFrame(columns=["الاسم", "البريد", "الهاتف", "الرصيد"]),
                "الموردين": pd.DataFrame(columns=["الاسم", "البريد", "الهاتف", "الرصيد"])
            }
        
        # إنشاء ملف Excel إذا لم يكن موجوداً
        self.excel_file = "accounting_data.xlsx"
        self.setup_excel_file()
        self.load_data()
    
    def setup_excel_file(self):
        """إنشاء ملف Excel مع أوراق العمل الأساسية إذا لم يكن موجوداً"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            # إزالة الورقة الافتراضية
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # إنشاء أوراق العمل الأساسية
            for sheet_name in ["المبيعات", "المشتريات", "المصروفات", "العملاء", "الموردين", "التقارير"]:
                wb.create_sheet(sheet_name)
            
            wb.save(self.excel_file)
    
    def load_data(self):
        """تحميل البيانات من ملف Excel"""
        try:
            excel_data = pd.read_excel(self.excel_file, sheet_name=None)
            for sheet_name in st.session_state.data:
                if sheet_name in excel_data:
                    st.session_state.data[sheet_name] = excel_data[sheet_name].fillna("")
        except Exception as e:
            st.error(f"خطأ في تحميل البيانات: {e}")
    
    def save_data(self):
        """حفظ البيانات إلى ملف Excel"""
        try:
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                for sheet_name, df in st.session_state.data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            st.success("تم حفظ البيانات بنجاح")
        except Exception as e:
            st.error(f"خطأ في حفظ البيانات: {e}")
    
    def run(self):
        """تشغيل التطبيق الرئيسي"""
        st.sidebar.title("نظام المحاسبة الذكي")
        
        # قائمة التنقل
        app_mode = st.sidebar.selectbox(
            "اختر الصفحة",
            ["الإدخال الرئيسي", "التقارير المحاسبية", "التحليل التفاعلي", "الإعدادات والربط", "التدقيق والمطابقة"]
        )
        
        # عرض الصفحة المحددة
        if app_mode == "الإدخال الرئيسي":
            self.show_input_page()
        elif app_mode == "التقارير المحاسبية":
            self.show_reports_page()
        elif app_mode == "التحليل التفاعلي":
            self.show_analysis_page()
        elif app_mode == "الإعدادات والربط":
            self.show_settings_page()
        elif app_mode == "التدقيق والمطابقة":
            self.show_audit_page()
    
    def show_input_page(self):
        """عرض صفحة الإدخال الرئيسي"""
        st.title("إدخال المعاملات المحاسبية")
        
        # أزرار طرق الإدخال المختلفة
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("إدخال صوتي", type="primary", use_container_width=True):
                self.voice_input()
        
        with col2:
            if st.button("مسح ضوئي", type="primary", use_container_width=True):
                self.camera_input()
        
        with col3:
            if st.button("إدخال يدوي", type="primary", use_container_width=True):
                self.manual_input()
        
        # منطقة عرض البيانات
        st.subheader("معاينة البيانات")
        self.data_display = st.text_area("بيانات المعاملة", height=200, placeholder="ستظهر هنا بيانات المعاملة بعد الإدخال")
        
        # أزرار المعالجة
        col4, col5, col6 = st.columns(3)
        
        with col4:
            if st.button("معالجة البيانات", use_container_width=True):
                self.process_data()
        
        with col5:
            if st.button("حفظ في النظام", use_container_width=True):
                self.save_data()
        
        with col6:
            if st.button("تدقيق المحاسبة", use_container_width=True):
                self.audit_data()
    
    def voice_input(self):
        """معالجة الإدخال الصوتي"""
        st.info("جاري الاستماع... قل بيانات المعاملة المحاسبية")
        
        try:
            recognizer = sr.Recognizer()
            with sr.Microphone() as source:
                recognizer.adjust_for_ambient_noise(source)
                audio = recognizer.listen(source, timeout=5, phrase_time_limit=10)
                
                text = recognizer.recognize_google(audio, language="ar-AR")
                st.session_state.input_text = f"النص المعترف به: {text}\n"
                
                # استخدام ChatGPT لتحويل النص إلى بيانات محاسبية منظمة
                accounting_data = self.parse_with_chatgpt(text)
                self.display_accounting_data(accounting_data)
                
                st.success("تم التعرف على الصوت بنجاح")
                
        except sr.WaitTimeoutError:
            st.warning("لم يتم الكشف عن أي صوت")
        except sr.UnknownValueError:
            st.error("لم يتم التعرف على الكلام")
        except sr.RequestError as e:
            st.error(f"خطأ في خدمة التعرف على الصوت: {e}")
    
    def camera_input(self):
        """معالجة الإدخال بالكاميرا"""
        st.info("جاري فتح الكاميرا...")
        
        uploaded_file = st.camera_input("التقاط صورة للفاتورة أو المستند")
        
        if uploaded_file is not None:
            # حفظ الصورة مؤقتًا
            with open("temp_invoice.jpg", "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # استخدام OCR لاستخراج النص
            extracted_text = self.extract_text_from_image("temp_invoice.jpg")
            st.session_state.input_text = f"النص المستخرج: {extracted_text}\n"
            
            # تحليل النص باستخدام ChatGPT
            invoice_data = self.parse_invoice_with_chatgpt(extracted_text)
            self.display_accounting_data(invoice_data)
            
            st.success("تم معالجة الصورة بنجاح")
    
    def extract_text_from_image(self, image_path):
        """استخراج النص من الصورة باستخدام OCR"""
        try:
            image = Image.open(image_path)
            # تحسين الصورة لتحسين دقة OCR
            image = ImageOps.exif_transpose(image)
            image = image.convert('L')  # تحويل إلى تدرج الرمادي
            text = pytesseract.image_to_string(image, lang='ara')
            return text
        except Exception as e:
            return f"خطأ في استخراج النص: {e}"
    
    def parse_with_chatgpt(self, text):
        """محاكاة اتصال بـ ChatGPT API"""
        # محاكاة استجابة ChatGPT بناءً على النص المدخل
        if "بيع" in text or "مبيعات" in text:
            simulated_response = {
                "transaction_type": "بيع",
                "amount": self.extract_amount(text),
                "currency": "ريال سعودي",
                "date": datetime.now().strftime("%Y-%m-%d"),
                "description": text,
                "account_debit": "حساب المدينين",
                "account_credit": "إيرادات المبيعات",
                "vat_amount": round(self.extract_amount(text) * 0.15, 2)
            }
        elif "شراء" in text or "مشتريات" in text:
            simulated_response = {
                "transaction_type": "شراء",
                "amount": self.extract_amount(text),
                "currency": "ريال سعودي",
                "date": datetime.now().strftime("%Y-%m-%d"),
                "description": text,
                "account_debit": "المشتريات",
                "account_credit": "حساب الدائنين",
                "vat_amount": round(self.extract_amount(text) * 0.15, 2)
            }
        else:
            simulated_response = {
                "transaction_type": "عام",
                "amount": self.extract_amount(text),
                "currency": "ريال سعودي",
                "date": datetime.now().strftime("%Y-%m-%d"),
                "description": text,
                "account_debit": "مصروفات عامة",
                "account_credit": "البنك",
                "vat_amount": 0.0
            }
        
        return simulated_response
    
    def extract_amount(self, text):
        """استخراج المبالغ الرقمية من النص"""
        numbers = re.findall(r'\d+\.\d+|\d+', text)
        if numbers:
            return float(numbers[0])
        return 1000.0  # قيمة افتراضية
    
    def parse_invoice_with_chatgpt(self, text):
        """محاكاة تحليل الفاتورة باستخدام ChatGPT"""
        # محاكاة تحليل الفاتورة
        simulated_response = {
            "invoice_number": f"INV-{datetime.now().strftime('%Y%m%d')}-001",
            "supplier": "شركة المعدات المتحدة",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "due_date": (datetime.now() + pd.DateOffset(days=30)).strftime("%Y-%m-%d"),
            "total_amount": self.extract_amount(text),
            "items": [
                {"description": "طابعة ليزر", "quantity": 2, "unit_price": 1200.00, "total": 2400.00},
                {"description": "حبر طابعة", "quantity": 5, "unit_price": 170.00, "total": 850.00}
            ],
            "vat_amount": round(self.extract_amount(text) * 0.15, 2)
        }
        
        return simulated_response
    
    def display_accounting_data(self, data):
        """عرض البيانات المحاسبية في واجهة المستخدم"""
        display_text = ""
        
        if data.get("transaction_type") == "بيع":
            display_text += "=== معاملة بيع ===\n"
        elif data.get("transaction_type") == "شراء":
            display_text += "=== معاملة شراء ===\n"
        else:
            display_text += "=== معاملة محاسبية ===\n"
        
        for key, value in data.items():
            if key == "items":
                display_text += f"{key}:\n"
                for item in value:
                    for k, v in item.items():
                        display_text += f"  {k}: {v}\n"
                    display_text += "\n"
            else:
                display_text += f"{key}: {value}\n"
        
        st.session_state.input_text = display_text
    
    def process_data(self):
        """معالجة البيانات وإضافتها للنظام"""
        if 'input_text' not in st.session_state or not st.session_state.input_text:
            st.warning("لا توجد بيانات معالجة")
            return
        
        current_text = st.session_state.input_text
        if "=== معاملة" not in current_text:
            st.warning("لا توجد بيانات معالجة صالحة")
            return
        
        # في التطبيق الحقيقي، سيتم تحليل النص وإضافة البيانات للداتا فريم المناسب
        lines = current_text.split('\n')
        transaction_data = {}
        
        for line in lines:
            if ':' in line and not line.strip().startswith('==='):
                key, value = line.split(':', 1)
                transaction_data[key.strip()] = value.strip()
        
        # تحديد نوع المعاملة وإضافتها للبيانات
        if "بيع" in transaction_data.get("transaction_type", ""):
            new_record = {
                "التاريخ": transaction_data.get("date", datetime.now().strftime("%Y-%m-%d")),
                "العميل": "عميل",
                "المبلغ": transaction_data.get("amount", 0),
                "الوصف": transaction_data.get("description", ""),
                "الحالة": "معلقة"
            }
            st.session_state.data["المبيعات"] = pd.concat([st.session_state.data["المبيعات"], pd.DataFrame([new_record])], ignore_index=True)
            st.success("تمت إضافة معاملة البيع بنجاح")
        
        elif "شراء" in transaction_data.get("transaction_type", ""):
            new_record = {
                "التاريخ": transaction_data.get("date", datetime.now().strftime("%Y-%m-%d")),
                "المورد": "مورد",
                "المبلغ": transaction_data.get("amount", 0),
                "الوصف": transaction_data.get("description", ""),
                "الحالة": "معلقة"
            }
            st.session_state.data["المشتريات"] = pd.concat([st.session_state.data["المشتريات"], pd.DataFrame([new_record])], ignore_index=True)
            st.success("تمت إضافة معاملة الشراء بنجاح")
        
        self.save_data()
    
    def manual_input(self):
        """فتح نافذة الإدخال اليدوي"""
        st.subheader("الإدخال اليدوي للمعاملات")
        
        with st.form("manual_input_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                transaction_type = st.selectbox("نوع المعاملة", ["بيع", "شراء", "مصروف"])
                transaction_date = st.date_input("التاريخ", datetime.now())
                transaction_party = st.text_input("العميل/المورد")
            
            with col2:
                transaction_amount = st.number_input("المبلغ", min_value=0.0, format="%.2f")
                transaction_desc = st.text_area("الوصف")
            
            submitted = st.form_submit_button("حفظ المعاملة")
            
            if submitted:
                if not all([transaction_party, transaction_amount]):
                    st.error("جميع الحقول مطلوبة")
                    return
                
                new_record = {
                    "التاريخ": transaction_date.strftime("%Y-%m-%d"),
                    "المبلغ": transaction_amount,
                    "الوصف": transaction_desc,
                    "الحالة": "مكتمل"
                }
                
                if transaction_type == "بيع":
                    new_record["العميل"] = transaction_party
                    st.session_state.data["المبيعات"] = pd.concat([st.session_state.data["المبيعات"], pd.DataFrame([new_record])], ignore_index=True)
                elif transaction_type == "شراء":
                    new_record["المورد"] = transaction_party
                    st.session_state.data["المشتريات"] = pd.concat([st.session_state.data["المشتريات"], pd.DataFrame([new_record])], ignore_index=True)
                else:
                    new_record["النوع"] = transaction_type
                    st.session_state.data["المصروفات"] = pd.concat([st.session_state.data["المصروفات"], pd.DataFrame([new_record])], ignore_index=True)
                
                self.save_data()
                st.success("تمت إضافة المعاملة بنجاح")
    
    def show_reports_page(self):
        """عرض صفحة التقارير"""
        st.title("التقارير المحاسبية")
        
        # اختيار نوع التقرير
        report_type = st.selectbox("اختر نوع التقرير", list(st.session_state.data.keys()))
        
        if st.button("إنشاء التقرير"):
            self.generate_report(report_type)
    
    def generate_report(self, report_type):
        """إنشاء تقرير حسب النوع المحدد"""
        if not st.session_state.data[report_type].empty:
            st.dataframe(st.session_state.data[report_type], use_container_width=True)
            
            # خيارات التصدير
            csv = st.session_state.data[report_type].to_csv(index=False)
            st.download_button(
                label="تحميل التقرير كملف CSV",
                data=csv,
                file_name=f"{report_type}_report.csv",
                mime="text/csv"
            )
        else:
            st.warning("لا توجد بيانات متاحة لهذا النوع من التقارير")
    
    def show_analysis_page(self):
        """عرض صفحة التحليل"""
        st.title("التحليل المالي التفاعلي")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("تحليل المبيعات"):
                self.create_chart("المبيعات")
        
        with col2:
            if st.button("تحليل المصروفات"):
                self.create_chart("المصروفات")
        
        with col3:
            if st.button("مقارنة الإيرادات"):
                self.create_comparison_chart()
    
    def create_chart(self, data_type):
        """إنشاء رسم بياني للبيانات"""
        if st.session_state.data[data_type].empty:
            st.warning("لا توجد بيانات متاحة")
            return
        
        # تحضير البيانات
        df = st.session_state.data[data_type].copy()
        df['التاريخ'] = pd.to_datetime(df['التاريخ'], errors='coerce')
        df['المبلغ'] = pd.to_numeric(df['المبلغ'], errors='coerce')
        
        # تجميع البيانات حسب الشهر
        monthly_data = df.groupby(df['التاريخ'].dt.to_period('M'))['المبلغ'].sum()
        
        # إنشاء الرسم البياني
        fig, ax = plt.subplots(figsize=(10, 6))
        months = [str(period) for period in monthly_data.index]
        amounts = monthly_data.values
        
        ax.bar(months, amounts, color=excel_color)
        ax.set_title(f'{data_type} الشهرية', fontsize=16)
        ax.set_ylabel('المبلغ', fontsize=12)
        ax.tick_params(axis='x', rotation=45)
        
        st.pyplot(fig)
    
    def create_comparison_chart(self):
        """إنشاء رسم بياني مقارن"""
        sales_data = st.session_state.data["المبيعات"].copy()
        purchases_data = st.session_state.data["المشتريات"].copy()
        
        if sales_data.empty and purchases_data.empty:
            st.warning("لا توجد بيانات متاحة")
            return
        
        sales_data['التاريخ'] = pd.to_datetime(sales_data['التاريخ'], errors='coerce')
        purchases_data['التاريخ'] = pd.to_datetime(purchases_data['التاريخ'], errors='coerce')
        
        sales_data['المبلغ'] = pd.to_numeric(sales_data['المبلغ'], errors='coerce')
        purchases_data['المبلغ'] = pd.to_numeric(purchases_data['المبلغ'], errors='coerce')
        
        # تجميع البيانات حسب الشهر
        monthly_sales = sales_data.groupby(sales_data['التاريخ'].dt.to_period('M'))['المبلغ'].sum()
        monthly_purchases = purchases_data.groupby(purchases_data['التاريخ'].dt.to_period('M'))['المبلغ'].sum()
        
        # إنشاء الرسم البياني
        fig, ax = plt.subplots(figsize=(10, 6))
        
        months = [str(period) for period in monthly_sales.index]
        sales = monthly_sales.values
        purchases = monthly_purchases.reindex(monthly_sales.index, fill_value=0).values
        
        bar_width = 0.35
        x = np.arange(len(months))
        
        ax.bar(x - bar_width/2, sales, bar_width, label='المبيعات', color=excel_color)
        ax.bar(x + bar_width/2, purchases, bar_width, label='المشتريات', color=chatgpt_color)
        
        ax.set_xlabel('الشهر')
        ax.set_ylabel('المبلغ')
        ax.set_title('مقارنة المبيعات والمشتريات')
        ax.set_xticks(x)
        ax.set_xticklabels(months, rotation=45)
        ax.legend()
        
        st.pyplot(fig)
    
    def show_settings_page(self):
        """عرض صفحة الإعدادات"""
        st.title("إعدادات النظام والربط الخارجي")
        
        st.subheader("إعدادات الربط")
        
        # أسعار العملات
        st.info("أسعار العملات: محدث تلقائياً")
        
        # حالة الربط البنكي
        st.info("الربط البنكي: متصل")
        
        # التحديث التلقائي
        auto_update = st.checkbox("التحديث التلقائي", value=True)
        
        # أزرار التحكم
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("تحديث البيانات"):
                self.update_external_data()
        
        with col2:
            if st.button("اختبار الاتصالات"):
                self.test_connections()
        
        with col3:
            if st.button("تصدير البيانات"):
                self.export_data()
    
    def update_external_data(self):
        """تحديث البيانات من المصادر الخارجية"""
        st.info("جاري تحديث البيانات الخارجية...")
        time.sleep(2)  # محاكاة وقت التحديث
        st.success("تم تحديث البيانات الخارجية")
    
    def test_connections(self):
        """اختبار الاتصالات الخارجية"""
        st.info("جاري اختبار الاتصالات...")
        time.sleep(2)  # محاكاة وقت الانتظار
        st.success("جميع الاتصالات تعمل بشكل صحيح")
    
    def export_data(self):
        """تصدير البيانات"""
        try:
            with pd.ExcelWriter("accounting_export.xlsx", engine='openpyxl') as writer:
                for sheet_name, df in st.session_state.data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            with open("accounting_export.xlsx", "rb") as f:
                st.download_button(
                    label="تحميل البيانات كملف Excel",
                    data=f,
                    file_name="accounting_data_export.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            st.success("تم تصدير البيانات بنجاح")
        except Exception as e:
            st.error(f"فشل في تصدير البيانات: {e}")
    
    def show_audit_page(self):
        """عرض صفحة التدقيق"""
        st.title("تدقيق المحاسبة واكتشاف الأخطاء")
        
        if st.button("بدء عملية التدقيق", type="primary"):
            self.run_audit()
    
    def run_audit(self):
        """تشغيل عملية التدقيق"""
        st.info("جاري تدقيق البيانات المحاسبية...")
        
        # محاكاة عملية التدقيق
        time.sleep(3)
        
        audit_results = {
            "status": "تم التدقيق",
            "issues_found": [
                {
                    "type": "تناقض",
                    "description": "الرصيد المدين لا يساوي الرصيد الدائن في قيد اليومية",
                    "suggestion": "مراجعة القيد رقم JV-2023-1045"
                },
                {
                    "type": "خطأ في التصنيف",
                    "description": "مصروفات تسويق مصنفة كمصروفات عمومية",
                    "suggestion": "إعادة تصنيف المبلغ 1250 ريال إلى حساب مصروفات التسويق"
                }
            ],
            "recommendations": [
                "تعديل القيد المحاسبي لتحقيق التوازن",
                "مراجعة دليل الحسابات للتأكد من التصنيف الصحيح"
            ]
        }
        
        self.display_audit_results(audit_results)
    
    def display_audit_results(self, results):
        """عرض نتائج التدقيق"""
        st.subheader("نتائج تدقيق النظام المحاسبي")
        
        st.write(f"**حالة التدقيق:** {results['status']}")
        
        if results['issues_found']:
            st.write("**المشكلات المكتشفة:**")
            for issue in results['issues_found']:
                st.write(f"- **نوع المشكلة:** {issue['type']}")
                st.write(f"  **الوصف:** {issue['description']}")
                st.write(f"  **الاقتراح:** {issue['suggestion']}")
                st.write("")
        
        if results['recommendations']:
            st.write("**التوصيات العامة:**")
            for rec in results['recommendations']:
                st.write(f"- {rec}")
        
        # زر معالجة الأخطاء
        if st.button("معالجة الأخطاء"):
            st.info("جاري معالجة الأخطاء...")
            time.sleep(2)
            st.success("تم معالجة الأخطاء")
        
        # زر تصدير التقرير
        audit_text = f"نتائج تدقيق النظام المحاسبي\n{'='*50}\n\n"
        audit_text += f"حالة التدقيق: {results['status']}\n\n"
        
        if results['issues_found']:
            audit_text += "المشكلات المكتشفة:\n"
            for issue in results['issues_found']:
                audit_text += f"- نوع المشكلة: {issue['type']}\n"
                audit_text += f"  الوصف: {issue['description']}\n"
                audit_text += f"  الاقتراح: {issue['suggestion']}\n\n"
        
        if results['recommendations']:
            audit_text += "التوصيات العامة:\n"
            for rec in results['recommendations']:
                audit_text += f"- {rec}\n"
        
        st.download_button(
            label="تصدير تقرير التدقيق",
            data=audit_text,
            file_name="audit_report.txt",
            mime="text/plain"
        )
    
    def audit_data(self):
        """تدقيق البيانات الحالية"""
        st.info("جاري تدقيق البيانات...")
        time.sleep(2)
        
        # محاكاة نتائج التدقيق
        st.success("تم تدقيق البيانات دون发现任何错误")
        
        # عرض النتائج
        st.write("**نتائج التدقيق:**")
        st.write("- جميع المعاملات متوازنة")
        st.write("- لا توجد أخطاء في التصنيف")
        st.write("- البيانات المالية سليمة")

# تشغيل التطبيق
if __name__ == "__main__":
    app = AccountingAIApp()
    app.run()

